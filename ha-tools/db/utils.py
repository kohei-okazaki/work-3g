from pathlib import Path
from jinja2 import Environment, FileSystemLoader
from openpyxl import load_workbook

"""
自動生成共通モジュール
"""


def get_template(template_file_name: str):
    """
    テンプレート環境を取得する関数
    @param template_file_name: テンプレートファイル名
    @return: Jinja2テンプレート環境
    """

    # Jinja2環境設定
    env = Environment(loader=FileSystemLoader("templates"), autoescape=False)

    # テンプレートファイルを取得
    return env.get_template(template_file_name)


def get_sheet_names(excel_path: str) -> list[str]:
    """
    Excelファイルからシート名リストを取得する関数
    @param excel_path: Excelファイルパス
    @return: シート名のリスト
    """

    # Excelファイル取得
    workbook = load_workbook(filename=Path(excel_path), data_only=True, read_only=True)

    # シート名リスト取得
    sheet_names: list[str] = workbook.sheetnames

    # Excelファイルを閉じる
    workbook.close()

    return sheet_names


def read_rows(
    excel_path: str,
    sheet_name: str,
    ignore_header: bool = False,
) -> list[str]:
    """
    Excelファイルから行情報を読み込む関数
    @param excel_path: Excelファイルパス
    @param sheet_name: Excelシート名
    @param ignore_header: ヘッダ行を無視するかどうか。無視する場合、2行目から読み込む。
    @return: 行情報のリスト
    """

    # Excelファイル取得
    workbook = load_workbook(filename=Path(excel_path), data_only=True, read_only=True)

    if sheet_name not in workbook.sheetnames:
        raise ValueError("{}シートが存在しません".format(sheet_name))

    # 対象シートを取得
    worksheet = workbook[sheet_name]

    # 最小行番号設定
    min_row: int = 2 if ignore_header else 1

    rows: list = []
    for row in worksheet.iter_rows(min_row=min_row, values_only=True):
        if row is None or all(cell is None for cell in row):
            # 空行はスキップ
            continue
        rows.append(row)

    # Excelファイルを閉じる
    workbook.close()

    return rows
