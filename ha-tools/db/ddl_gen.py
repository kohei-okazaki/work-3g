from pathlib import Path
from openpyxl import load_workbook
from jinja2 import Environment, FileSystemLoader

"""
DDLを自動生成するモジュール
"""

# DB定義書パス
EXCEL_PATH = r"C:\app\git\work-3g\ha-asset\02_db\DB.xlsx"
# DDL出力パス
OUTPUT_PATH = r"C:\app\git\work-3g\ha-asset\02_db\ddl\test"
# テーブル名（論理）定義位置
LOGICAL_TABLE_NAME_POS = 0
# テーブル名（物理）定義位置
PHYSICAL_TABLE_NAME_POS = 1
# PKフラグ定義位置
PK_FLAG_POS = 2
# シーケンスフラグ定義位置
SEQUENCE_FLAG_POS = 3
# 暗号化フラグ定義位置
CRYPT_FLAG_POS = 4
# NotNullフラグ定義位置
NOTNULL_FLAG_POS = 5
# カラム名（論理）定義位置
LOGICAL_COLUMN_NAME_POS = 6
# カラム名（物理）定義位置
PHYSICAL_COLUMN_NAME_POS = 7
# データ型定義位置
DATA_TYPE_POS = 8
# サイズ定義位置
SIZE_POS = 9
# デフォルト値定義位置
DEFAULT_VALUE_POS = 10


def execute():
    """
    DDL自動生成のメイン関数
    """
    tables: dict[str, list[dict]] = read_tables()
    # print(tables)
    generate_ddl(tables)


def read_tables() -> dict[str, list[dict]]:
    """
    Excelファイルからテーブルリストを読み込む関数
    @return: (テーブル名, カラム情報リスト)の辞書
    """

    # Excelファイル取得
    workbook = load_workbook(filename=Path(EXCEL_PATH), data_only=True, read_only=True)

    if "TABLE_LIST" not in workbook.sheetnames:
        raise ValueError("{}シートが存在しません".format("TABLE_LIST"))

    # TABLE_LISTシートを取得
    worksheet = workbook["TABLE_LIST"]

    rows: list = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if row is None or all(cell is None for cell in row):
            # 空行はスキップ
            continue
        rows.append(row)

    # Excelファイルを閉じる
    workbook.close()

    # テーブル名->カラム情報リストのdict
    tables: dict[str, list[dict]] = {}

    for row in rows:
        # テーブル名（論理）
        logical_table_name = row[LOGICAL_TABLE_NAME_POS]
        # テーブル名（物理）
        physical_table_name = row[PHYSICAL_TABLE_NAME_POS]
        # PKフラグ
        is_primary_key = row[PK_FLAG_POS] is not None and row[PK_FLAG_POS] == "1"
        # シーケンスフラグ
        is_sequence = (
            row[SEQUENCE_FLAG_POS] is not None and row[SEQUENCE_FLAG_POS] == "1"
        )
        # 暗号化フラグ
        is_crypt = row[CRYPT_FLAG_POS] is not None and row[CRYPT_FLAG_POS] == "1"
        # NotNullフラグ
        is_not_null = row[NOTNULL_FLAG_POS] is not None and row[NOTNULL_FLAG_POS] == "1"
        # カラム名(論理)
        logical_column_name = row[LOGICAL_COLUMN_NAME_POS]
        # カラム名(物理)
        physical_column_name = row[PHYSICAL_COLUMN_NAME_POS]
        # データ型
        data_type = row[DATA_TYPE_POS]
        # サイズ
        size = row[SIZE_POS]
        # データ型にサイズを付与
        if size is not None and size != "":
            # サイズ指定がある場合
            data_type = data_type + "({})".format(size)
        # デフォルト値
        default_value = row[DEFAULT_VALUE_POS]
        if (
            row[DATA_TYPE_POS] is not None
            and row[DATA_TYPE_POS] in ("VARCHAR", "DATE")
            and default_value is not None
        ):
            # デフォルト値が文字列型 or 日付型の場合、シングルクォートで囲む
            default_value = "'{}'".format(default_value)

        # column_info辞書を作成
        column_info = {
            "logical_column_name": logical_column_name,
            "physical_column_name": physical_column_name,
            "is_primary_key": is_primary_key,
            "is_sequence": is_sequence,
            "is_crypt": is_crypt,
            "is_not_null": is_not_null,
            "type": data_type,
            "has_default": default_value is not None and default_value != "",
            "default_value": default_value,
        }

        # テーブル名(論理名.物理名)をキーにしてカラム情報を追加
        table_key = "{}.{}".format(logical_table_name, physical_table_name)
        if table_key in tables:
            # テーブルが存在する場合、既存テーブルにカラム情報を追加
            tables[table_key].append(column_info)
        else:
            # テーブルが存在しない場合、新規作成してカラム情報を追加
            tables[table_key] = [column_info]

    return tables


def generate_ddl(tables: dict[str, list[dict]]):
    """
    DDLを生成してファイルに出力する
    @param tables: (テーブル名, カラム情報リスト)の辞書
    """

    env = Environment(loader=FileSystemLoader("templates"), autoescape=False)

    ddl_template = env.get_template("ddl_template.sql.j2")

    # 出力ファイルパスを作成
    output_file_path = Path(OUTPUT_PATH)
    output_file_path.mkdir(exist_ok=True)

    for table_name, columns in tables.items():

        # テーブル名（論理）
        logical_table_name = table_name.split(".")[LOGICAL_TABLE_NAME_POS]
        # テーブル名（物理）
        physical_table_name = table_name.split(".")[PHYSICAL_TABLE_NAME_POS]

        ddl = ddl_template.render(
            logical_table_name=logical_table_name,
            physical_table_name=physical_table_name,
            columns=columns,
        )

        ddl_file_name = "{}.sql".format(physical_table_name)
        (output_file_path / ddl_file_name).write_text(ddl, encoding="utf-8")


if __name__ == "__main__":
    execute()
