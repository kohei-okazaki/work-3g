from pathlib import Path
from openpyxl import load_workbook

"""
自動生成共通モジュール
"""


def read_rows(
    excel_path: str,
    sheet_name: str,
) -> list[str]:
    """
    Excelファイルから行情報を読み込む関数
    @param excel_path: Excelファイルパス
    @param sheet_name: Excelシート名
    @return: 行情報のリスト
    """

    # Excelファイル取得
    workbook = load_workbook(filename=Path(excel_path), data_only=True, read_only=True)

    if sheet_name not in workbook.sheetnames:
        raise ValueError("{}シートが存在しません".format(sheet_name))

    # 対象シートを取得
    worksheet = workbook[sheet_name]

    rows: list = []
    for row in worksheet.iter_rows(min_row=1, values_only=True):
        if row is None or all(cell is None for cell in row):
            # 空行はスキップ
            continue
        rows.append(row)

    # Excelファイルを閉じる
    workbook.close()

    return rows
